"""Generate test cases per Jira ticket — Excel format like AI Control Plane."""

import io
import os
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()

from core.llm_gateway import gateway

DARK_BLUE = "1F4E78"
LIGHT_BLUE = "D9E1F2"
GREEN_LIGHT = "E8F5E9"
WHITE = "FFFFFF"


def generate_test_cases_for_story(story: dict, epic: dict = None) -> list:
    """Generate test cases using the anti-hallucination generator from AI Control Plane."""
    from services.test_case_generator import generate_test_cases, story_to_jira_context

    ctx = story_to_jira_context(story, epic=epic, project="SDLC")
    result = generate_test_cases(ctx)
    suite = result.get("test_suite", []) or []

    # Adapt the AI Control Plane output shape to your existing Excel writer's expectations.
    # Excel writer expects: tc_id, title, type, preconditions, steps, expected_result, priority
    adapted = []
    for tc in suite:
        cat = (tc.get("category") or "").lower()
        # Map AI Control Plane categories → your existing color/type mapping
        if cat in ("negative", "edge case"):
            t = "negative" if cat == "negative" else "edge"
        elif cat in ("functional", "integration", "regression", "reproduction"):
            t = "positive"
        else:
            t = "positive"
        adapted.append({
            "tc_id": tc.get("id", ""),
            "title": tc.get("title", ""),
            "type": t,
            "preconditions": tc.get("preconditions", ""),
            "steps": tc.get("steps", []),
            "expected_result": tc.get("expected_result", ""),
            "priority": tc.get("priority", "P2"),
        })
    return adapted


def export_test_cases_excel(pipeline_state: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["H"].width = 10

    # Header
    title_cell = ws.cell(row=1, column=1, value=f"Test Cases — {pipeline_state.get('thread_id','')}")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", start_color=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ["TC ID", "Story / Ticket", "Test Title", "Type", "Preconditions",
               "Steps", "Expected Result", "Priority"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28

    state = pipeline_state.get("current_state", {})
    sprint_plan = state.get("sprint_plan", {})
    jira_tickets = state.get("jira_tickets", [])

    row = 4
    tc_counter = 1

    for epic in sprint_plan.get("epics", []):
        for story in epic.get("stories", []):
            story_title = story.get("title", "")
            ticket_key = ""
            # Try to match story to Jira ticket by index
            if jira_tickets and len(jira_tickets) > 0:
                # Story tickets are usually indexed after epic tickets
                pass

            test_cases = generate_test_cases_for_story(story, epic=epic)

            for tc in test_cases:
                tc_id = f"TC-{tc_counter:03d}"
                tc_counter += 1

                ws.cell(row=row, column=1, value=tc_id).font = Font(bold=True, color="2A7AE4")
                ws.cell(row=row, column=2, value=story_title)
                ws.cell(row=row, column=3, value=tc.get("title", ""))

                tc_type = tc.get("type", "positive")
                type_color = "1A6B3A" if tc_type == "positive" else "C0392B" if tc_type == "negative" else "B7860B"
                type_cell = ws.cell(row=row, column=4, value=tc_type.upper())
                type_cell.font = Font(bold=True, color=type_color)

                ws.cell(row=row, column=5, value=tc.get("preconditions", ""))
                steps = "\n".join([f"{i+1}. {s}" for i, s in enumerate(tc.get("steps", []))])
                ws.cell(row=row, column=6, value=steps)
                ws.cell(row=row, column=7, value=tc.get("expected_result", ""))

                pri = tc.get("priority", "P2")
                pri_color = "C0392B" if pri == "P1" else "B7860B" if pri == "P2" else "1A6B3A"
                pri_cell = ws.cell(row=row, column=8, value=pri)
                pri_cell.font = Font(bold=True, color=pri_color)

                # Wrap text on all
                for col in range(1, 9):
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

                ws.row_dimensions[row].height = max(40, len(steps) // 50 * 15 + 30)
                row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()