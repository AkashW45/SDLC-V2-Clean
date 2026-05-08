# api/runbook_export.py
# Enterprise Excel export — matches V1 runbook format
# Structure: header info → deployment matrix → group sections → test cases per ticket

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DARK_BLUE  = "1F4E78"
LIGHT_BLUE = "D9E1F2"
RED_LIGHT  = "FDECEA"
GREEN_LIGHT= "E8F5E9"
ORANGE_LIGHT="FFF3E0"
PURPLE_LIGHT="EDE7F6"
WHITE      = "FFFFFF"

GROUP_COLORS = {
    "migration":  ("FFF3E0", "B7860B"),
    "bugfix":     ("FDECEA", "C0392B"),
    "feature":    ("E8F5E9", "1A6B3A"),
    "deployment": ("D9E1F2", "1F4E78"),
    "testing":    ("EDE7F6", "4A235A"),
}


def _bold(ws, row, col, value, fill=None, size=11, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=size, bold=True, color=color)
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def _val(ws, row, col, value, wrap=True, size=11, color="000000", bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    c.font = Font(name="Calibri", size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    if fill:
        c.fill = PatternFill("solid", start_color=fill)
    return c


def _thdr(ws, row, col, value, bg=DARK_BLUE):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def _section(ws, row, title, ncols=9, bg=LIGHT_BLUE, text_color="000000"):
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Calibri", size=11, bold=True, color=text_color)
    c.fill = PatternFill("solid", start_color=bg)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 20
    return row + 1


def _kv(ws, row, key, value):
    _bold(ws, row, 1, key)
    _val(ws, row, 2, value)
    return row + 1


def _numbered_list(ws, row, items):
    for i, item in enumerate(items, 1):
        _val(ws, row, 1, f"{i}.", bold=True)
        c = ws.cell(row=row, column=2, value=str(item))
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(30, len(str(item))//60 * 15 + 15)
        row += 1
    return row


def export_runbook_excel(pipeline_state: dict) -> bytes:
    """
    Build enterprise-grade Excel runbook from full pipeline state.
    Includes:
      - Header (project, version, date)
      - Deployment matrix (per repo)
      - Runbook summary + pre-checks + global steps
      - Group sections (migration/bugfix/feature/deployment) — per group rollback
      - Validation + Rollback + Escalation
      - Sprint tickets (Jira)
      - Test cases (if available)
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Release Runbook"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 18

    # Get state
    requirement = pipeline_state.get("requirement", "")
    state       = pipeline_state.get("current_state", {})
    brd         = state.get("brd", {})
    prd         = state.get("prd", {})
    arch        = state.get("architecture", {})
    sprint_plan = state.get("sprint_plan", {})
    runbook     = state.get("runbook", {})
    impact      = state.get("impact_report", {})
    jira_tickets= state.get("jira_tickets", [])
    pr_urls     = pipeline_state.get("pr_urls", [])

    row = 1

    # ── Header ────────────────────────────────────────────
    row = _kv(ws, row, "Change ID",          pipeline_state.get("thread_id", ""))
    row = _kv(ws, row, "Release Description", brd.get("title") or prd.get("title") or requirement[:80])
    row = _kv(ws, row, "Project",             brd.get("title", "SDLC-V2"))
    row = _kv(ws, row, "Risk Level",          impact.get("risk_assessment", {}).get("risk_level", "TBD").upper())
    row = _kv(ws, row, "Generated",           datetime.now().strftime("%Y-%m-%d %H:%M"))
    row = _kv(ws, row, "Pipeline Status",     pipeline_state.get("status", ""))
    row += 1

    # ── Requirement Block ─────────────────────────────────
    row = _section(ws, row, "Original Requirement")
    c = ws.cell(row=row, column=1, value=requirement)
    c.font = Font(name="Calibri", size=11)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.row_dimensions[row].height = max(30, len(requirement)//100 * 15 + 15)
    row += 2

    # ── BRD Summary ────────────────────────────────────────
    if brd:
        row = _section(ws, row, "Business Requirements (BRD)")
        if brd.get("business_objectives"):
            row = _kv(ws, row, "Business Objectives",
                      "\n".join([f"• {o}" for o in brd["business_objectives"]]))
            ws.row_dimensions[row-1].height = max(30, len(brd["business_objectives"]) * 15)
        if brd.get("scope", {}).get("in_scope"):
            row = _kv(ws, row, "In Scope",
                      "\n".join([f"• {s}" for s in brd["scope"]["in_scope"]]))
            ws.row_dimensions[row-1].height = max(30, len(brd["scope"]["in_scope"]) * 15)
        if brd.get("risks"):
            row = _kv(ws, row, "Risks",
                      "\n".join([f"• {r}" for r in brd["risks"]]))
            ws.row_dimensions[row-1].height = max(30, len(brd["risks"]) * 15)
        row += 1

    # ── Architecture Components ────────────────────────────
    if arch and arch.get("nodes"):
        row = _section(ws, row, "Architecture Components")
        for col, hdr in enumerate(["ID","Name","Type","Zone","Tech Stack","Responsibilities"], start=1):
            _thdr(ws, row, col, hdr)
        ws.row_dimensions[row].height = 25
        row += 1
        for node in arch.get("nodes", []):
            _val(ws, row, 1, node.get("id", ""), bold=True, color="2A7AE4")
            _val(ws, row, 2, node.get("name", ""))
            _val(ws, row, 3, node.get("type", ""))
            _val(ws, row, 4, node.get("zone", ""))
            _val(ws, row, 5, ", ".join(node.get("tech_stack", [])))
            _val(ws, row, 6, "\n".join([f"• {r}" for r in node.get("responsibilities", [])]))
            ws.row_dimensions[row].height = max(35, len(node.get("responsibilities", [])) * 15)
            row += 1
        row += 1

    # ── Deployment Matrix ──────────────────────────────────
    affected_repos = impact.get("affected_repos", [])
    if affected_repos:
        row = _section(ws, row, "Deployment Matrix")
        for col, hdr in enumerate(
            ["Repo","Active Rail","Release Version","Rollback Version","PR Link","Remarks"],
            start=1
        ):
            _thdr(ws, row, col, hdr)
        ws.row_dimensions[row].height = 25
        row += 1

        for repo in affected_repos:
            _val(ws, row, 1, repo, bold=True)
            _val(ws, row, 2, "QA → PROD")
            _val(ws, row, 3, "v" + datetime.now().strftime("%Y.%m.%d"))
            _val(ws, row, 4, "previous tag")
            pr_for_repo = next((u for u in pr_urls if repo in u), "")
            _val(ws, row, 5, pr_for_repo or "TBD", color="2A7AE4")
            _val(ws, row, 6, "AI-generated changes")
            row += 1
        row += 1

    # ── Sprint Tickets ─────────────────────────────────────
    if jira_tickets:
        row = _section(ws, row, "Sprint Tickets (Jira)")
        for col, hdr in enumerate(["Ticket Key","Type","Status"], start=1):
            _thdr(ws, row, col, hdr)
        ws.row_dimensions[row].height = 25
        row += 1
        for tkt in jira_tickets:
            ticket_type = "Epic" if tkt.startswith("DEV-") and len(tkt) < 9 else "Story"
            _val(ws, row, 1, tkt, bold=True, color="2A7AE4")
            _val(ws, row, 2, ticket_type)
            _val(ws, row, 3, "Open")
            row += 1
        row += 1

    # ── Runbook Summary ────────────────────────────────────
    if runbook:
        row = _section(ws, row, "Runbook Summary")
        if runbook.get("feature"):
            row = _kv(ws, row, "Feature",  runbook["feature"])
        if runbook.get("version"):
            row = _kv(ws, row, "Version", runbook["version"])
        row += 1

        # Pre-deployment checklist
        if runbook.get("pre_deployment_checklist"):
            row = _section(ws, row, "Pre-Deployment Checklist")
            row = _numbered_list(ws, row, runbook["pre_deployment_checklist"])
            row += 1

        # Deployment sequence
        if runbook.get("deployment_sequence"):
            row = _section(ws, row, "Deployment Sequence")
            for col, hdr in enumerate(
                ["Step","Repo","Action","Command","Rollback Command"],
                start=1
            ):
                _thdr(ws, row, col, hdr)
            ws.row_dimensions[row].height = 25
            row += 1
            for s in runbook["deployment_sequence"]:
                _val(ws, row, 1, s.get("step", ""), bold=True)
                _val(ws, row, 2, s.get("repo", ""))
                _val(ws, row, 3, s.get("action", ""))
                cmd_cell = ws.cell(row=row, column=4, value=s.get("command", ""))
                cmd_cell.font = Font(name="Courier New", size=9, color="CDD6F4")
                cmd_cell.fill = PatternFill("solid", start_color="1E1E2E")
                cmd_cell.alignment = Alignment(wrap_text=True, vertical="top")
                rb_cell = ws.cell(row=row, column=5, value=s.get("rollback_command", ""))
                rb_cell.font = Font(name="Courier New", size=9, color="000000")
                rb_cell.fill = PatternFill("solid", start_color="FDECEA")
                rb_cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[row].height = 30
                row += 1
            row += 1

        # Feature flags
        if runbook.get("feature_flags"):
            row = _section(ws, row, "Feature Flags")
            for col, hdr in enumerate(
                ["Flag Name","Default","Enable After Deploy"], start=1
            ):
                _thdr(ws, row, col, hdr)
            ws.row_dimensions[row].height = 25
            row += 1
            for f in runbook["feature_flags"]:
                _val(ws, row, 1, f.get("flag_name", ""), bold=True, color="2A7AE4")
                _val(ws, row, 2, str(f.get("default", False)).upper())
                _val(ws, row, 3, str(f.get("enable_after_deploy", True)).upper(),
                     color="1A6B3A" if f.get("enable_after_deploy") else "C0392B")
                row += 1
            row += 1

        # Smoke test
        if runbook.get("smoke_test_checklist"):
            row = _section(ws, row, "Smoke Test Checklist", bg=GREEN_LIGHT)
            row = _numbered_list(ws, row, runbook["smoke_test_checklist"])
            row += 1

        # Rollback decision criteria
        if runbook.get("rollback_decision_criteria"):
            row = _section(ws, row, "Rollback Decision Criteria", bg=ORANGE_LIGHT)
            row = _numbered_list(ws, row, runbook["rollback_decision_criteria"])
            row += 1

        # Rollback steps
        if runbook.get("rollback_steps"):
            row = _section(ws, row, "Rollback Procedure", bg=RED_LIGHT)
            row = _numbered_list(ws, row, runbook["rollback_steps"])
            row += 1

        # On-call escalation
        oncall = runbook.get("on_call_escalation", {})
        if oncall:
            row = _section(ws, row, "On-Call Escalation")
            row = _kv(ws, row, "Primary",        oncall.get("primary", ""))
            row = _kv(ws, row, "Secondary",      oncall.get("secondary", ""))
            row = _kv(ws, row, "Slack Channel",  oncall.get("slack_channel", ""))
            row += 1

    # ── Impact Report ──────────────────────────────────────
    if impact:
        risk = impact.get("risk_assessment", {})
        row = _section(ws, row, "Impact Analysis")
        row = _kv(ws, row, "Risk Level",     risk.get("risk_level", "").upper())
        row = _kv(ws, row, "Recommendation", risk.get("recommendation", ""))
        if risk.get("breaking_changes"):
            row = _kv(ws, row, "Breaking Changes",
                      "\n".join([f"• {b}" for b in risk["breaking_changes"]]))
            ws.row_dimensions[row-1].height = max(30, len(risk["breaking_changes"]) * 15)
        if impact.get("affected_files"):
            row = _section(ws, row, "Affected Files")
            for col, hdr in enumerate(["File","Repo","Score"], start=1):
                _thdr(ws, row, col, hdr)
            ws.row_dimensions[row].height = 25
            row += 1
            for f in impact["affected_files"]:
                _val(ws, row, 1, f.get("file_path", ""))
                _val(ws, row, 2, f.get("repo_name", ""))
                _val(ws, row, 3, f.get("relevance_score", ""))
                row += 1
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_brd_markdown(brd: dict) -> str:
    if not brd:
        return "# BRD\n\n_No BRD generated yet._"

    L = []
    L.append(f"# Business Requirements Document")
    L.append(f"## {brd.get('title', 'Untitled')}\n")

    if brd.get("executive_summary"):
        L.append("## Executive Summary\n")
        L.append(brd["executive_summary"] + "\n")

    if brd.get("business_context"):
        L.append("## Business Context\n")
        L.append(brd["business_context"] + "\n")

    if brd.get("business_objectives"):
        L.append("## Business Objectives\n")
        for o in brd["business_objectives"]:
            L.append(f"- {o}")
        L.append("")

    if brd.get("in_scope"):
        L.append("## Scope\n\n### In Scope\n")
        for s in brd["in_scope"]:
            L.append(f"- {s}")
        L.append("")

    if brd.get("out_of_scope"):
        L.append("### Out of Scope\n")
        for s in brd["out_of_scope"]:
            L.append(f"- {s}")
        L.append("")

    if brd.get("stakeholders"):
        L.append("## Stakeholders\n")
        L.append("| Role | Name/Team | Responsibility |")
        L.append("|------|-----------|----------------|")
        for s in brd["stakeholders"]:
            if isinstance(s, dict):
                L.append(f"| {s.get('role','')} | {s.get('name_or_team','')} | {s.get('responsibility','')} |")
            else:
                L.append(f"| - | {s} | - |")
        L.append("")

    if brd.get("raci_matrix"):
        L.append("## RACI Matrix\n")
        L.append("| Activity | Responsible | Accountable | Consulted | Informed |")
        L.append("|----------|-------------|-------------|-----------|----------|")
        for r in brd["raci_matrix"]:
            L.append(f"| {r.get('activity','')} | {r.get('responsible','')} | {r.get('accountable','')} | {r.get('consulted','')} | {r.get('informed','')} |")
        L.append("")

    if brd.get("functional_requirements"):
        L.append(f"## Functional Requirements ({len(brd['functional_requirements'])})\n")
        for fr in brd["functional_requirements"]:
            L.append(f"### {fr.get('id','')} — {fr.get('title','')}")
            L.append(f"**Priority:** {fr.get('priority','')}\n")
            L.append(f"{fr.get('description','')}\n")
            if fr.get("business_value"):
                L.append(f"**Business Value:** {fr['business_value']}\n")

    if brd.get("non_functional_requirements"):
        L.append(f"## Non-Functional Requirements ({len(brd['non_functional_requirements'])})\n")
        L.append("| ID | Title | Description | Priority | Metric |")
        L.append("|-----|-------|-------------|----------|--------|")
        for nfr in brd["non_functional_requirements"]:
            L.append(f"| {nfr.get('id','')} | {nfr.get('title','')} | {nfr.get('description','')} | {nfr.get('priority','')} | {nfr.get('metric','')} |")
        L.append("")

    if brd.get("kpis"):
        L.append("## Key Performance Indicators\n")
        L.append("| KPI | Target | Method | Frequency |")
        L.append("|-----|--------|--------|-----------|")
        for k in brd["kpis"]:
            L.append(f"| {k.get('name','')} | {k.get('target','')} | {k.get('measurement_method','')} | {k.get('frequency','')} |")
        L.append("")

    if brd.get("risk_matrix"):
        L.append("## Risk Matrix\n")
        L.append("| ID | Risk | Likelihood | Impact | Mitigation | Owner |")
        L.append("|----|------|------------|--------|------------|-------|")
        for r in brd["risk_matrix"]:
            L.append(f"| {r.get('id','')} | {r.get('risk','')} | {r.get('likelihood','')} | {r.get('impact','')} | {r.get('mitigation','')} | {r.get('owner','')} |")
        L.append("")

    if brd.get("assumptions"):
        L.append("## Assumptions\n")
        for a in brd["assumptions"]:
            L.append(f"- {a}")
        L.append("")

    if brd.get("dependencies"):
        L.append("## Dependencies\n")
        for d in brd["dependencies"]:
            L.append(f"- {d}")
        L.append("")

    if brd.get("success_criteria"):
        L.append("## Success Criteria\n")
        for s in brd["success_criteria"]:
            L.append(f"- {s}")
        L.append("")

    if brd.get("timeline_estimate"):
        L.append(f"## Timeline\n\n{brd['timeline_estimate']}\n")

    if brd.get("budget_considerations"):
        L.append(f"## Budget Considerations\n\n{brd['budget_considerations']}\n")

    return "\n".join(L)

def export_prd_markdown(prd: dict) -> str:
    if not prd:
        return "# PRD\n\n_No PRD generated yet._"

    L = []
    L.append(f"# Product Requirements Document")
    L.append(f"## {prd.get('title', 'Untitled')}\n")

    if prd.get("executive_summary"):
        L.append("## Executive Summary\n")
        L.append(prd["executive_summary"] + "\n")

    if prd.get("product_vision"):
        L.append(f"## Product Vision\n\n{prd['product_vision']}\n")

    if prd.get("target_users"):
        L.append("## Target Users\n")
        for u in prd["target_users"]:
            if isinstance(u, dict):
                L.append(f"### {u.get('persona','')}")
                L.append(f"- **Needs:** {u.get('needs','')}")
                L.append(f"- **Pain Points:** {u.get('pain_points','')}\n")

    if prd.get("user_journeys"):
        L.append("## User Journeys\n")
        for j in prd["user_journeys"]:
            if isinstance(j, dict):
                L.append(f"### {j.get('journey','')}")
                for i, s in enumerate(j.get("steps", []), 1):
                    L.append(f"{i}. {s}")
                L.append("")

    if prd.get("functional_requirements"):
        L.append(f"## Functional Requirements ({len(prd['functional_requirements'])})\n")
        for fr in prd["functional_requirements"]:
            L.append(f"### {fr.get('id','')} — {fr.get('title','')}")
            L.append(f"**Priority:** {fr.get('priority','')}\n")
            if fr.get("user_story"):
                L.append(f"**User Story:** {fr['user_story']}\n")
            L.append(f"{fr.get('description','')}\n")
            if fr.get("acceptance_criteria"):
                L.append("**Acceptance Criteria:**")
                for ac in fr["acceptance_criteria"]:
                    L.append(f"- {ac}")
                L.append("")
            if fr.get("edge_cases"):
                L.append("**Edge Cases:**")
                for ec in fr["edge_cases"]:
                    L.append(f"- {ec}")
                L.append("")
            if fr.get("dependencies"):
                L.append(f"**Dependencies:** {', '.join(fr['dependencies'])}\n")

    if prd.get("non_functional_requirements"):
        L.append("## Non-Functional Requirements\n")
        L.append("| ID | Title | Description | Priority | Verification |")
        L.append("|-----|-------|-------------|----------|--------------|")
        for nfr in prd["non_functional_requirements"]:
            L.append(f"| {nfr.get('id','')} | {nfr.get('title','')} | {nfr.get('description','')} | {nfr.get('priority','')} | {nfr.get('verification_method','')} |")
        L.append("")

    if prd.get("technical_requirements"):
        L.append("## Technical Requirements\n")
        for tr in prd["technical_requirements"]:
            L.append(f"### {tr.get('id','')} — {tr.get('title','')}")
            L.append(f"{tr.get('description','')}\n")
            if tr.get("rationale"):
                L.append(f"**Rationale:** {tr['rationale']}\n")

    if prd.get("success_metrics"):
        L.append("## Success Metrics\n")
        L.append("| Metric | Baseline | Target | Timeline |")
        L.append("|--------|----------|--------|----------|")
        for m in prd["success_metrics"]:
            L.append(f"| {m.get('metric','')} | {m.get('baseline','')} | {m.get('target','')} | {m.get('timeline','')} |")
        L.append("")

    if prd.get("release_phases"):
        L.append("## Release Plan\n")
        for r in prd["release_phases"]:
            L.append(f"### {r.get('phase','')} — {r.get('timeline','')}")
            L.append(f"**Scope:** {', '.join(r.get('scope', []))}\n")

    if prd.get("open_questions"):
        L.append("## Open Questions\n")
        for q in prd["open_questions"]:
            L.append(f"- {q}")
        L.append("")

    return "\n".join(L)

def export_adr_markdown(adr: dict) -> str:
    if not adr:
        return "# ADR\n\n_No ADRs generated yet._"
    lines = ["# Architecture Decision Records\n"]
    for d in adr.get("decisions", []):
        lines.append(f"## {d.get('id','')} — {d.get('title','')}")
        lines.append(f"**Status:** {d.get('status','Accepted')}\n")
        lines.append(f"### Context\n\n{d.get('context','')}\n")
        lines.append(f"### Decision\n\n{d.get('decision','')}\n")
        if d.get("consequences"):
            lines.append("### Consequences\n")
            for c in d["consequences"]:
                lines.append(f"- {c}")
            lines.append("")
        if d.get("alternatives_considered"):
            lines.append("### Alternatives Considered\n")
            for a in d["alternatives_considered"]:
                lines.append(f"- {a}")
            lines.append("")
    return "\n".join(lines)


def export_architecture_markdown(architecture: dict) -> str:
    """Export architecture with embedded Mermaid diagram."""
    if not architecture or not architecture.get("nodes"):
        return "# Architecture\n\nNo architecture data available.\n"

    lines = []
    lines.append(f"# System Architecture: {architecture.get('system_name', 'Untitled System')}\n")
    lines.append(f"**Style:** {architecture.get('architecture_style', 'N/A')}")
    lines.append(f"**Deployment:** {architecture.get('deployment_model', 'N/A')}\n")

    # ── MERMAID DIAGRAM ──────────────────────────────────────────
    lines.append("## Architecture Diagram\n")
    mermaid_code = architecture.get("mermaid", "") or architecture.get("mermaid_diagram", "")
    if not mermaid_code:
        # Fallback: generate Mermaid from nodes/edges if not pre-generated
        mermaid_code = _generate_mermaid_from_nodes(architecture)

    lines.append("```mermaid")
    lines.append(mermaid_code)
    lines.append("```\n")

    # Components
    lines.append("## Components\n")
    for node in architecture.get("nodes", []):
        lines.append(f"### {node.get('id','')} — {node.get('name','')}")
        lines.append(f"- **Type:** {node.get('type','')}")
        lines.append(f"- **Zone:** {node.get('zone','')}")
        if node.get('tech_stack'):
            lines.append(f"- **Tech Stack:** {', '.join(node['tech_stack'])}")
        lines.append(f"- **Description:** {node.get('description','')}")
        if node.get('responsibilities'):
            lines.append("- **Responsibilities:**")
            for r in node['responsibilities']:
                lines.append(f"  - {r}")
        lines.append("")

    # Connections
    lines.append("## Connections\n")
    for edge in architecture.get("edges", []):
        proto = edge.get('protocol', '')
        desc = edge.get('description', '')
        lines.append(f"- `{edge.get('source','')}` → `{edge.get('target','')}` [{proto}] — {desc}")
    lines.append("")

    # Security
    if architecture.get("security_considerations"):
        lines.append("## Security Considerations\n")
        for s in architecture["security_considerations"]:
            lines.append(f"- {s}")
        lines.append("")

    # Scalability
    if architecture.get("scalability_notes"):
        lines.append("## Scalability\n")
        lines.append(architecture["scalability_notes"])
        lines.append("")

    return "\n".join(lines)


def _generate_mermaid_from_nodes(architecture: dict) -> str:
    """Generate Mermaid graph syntax from architecture nodes/edges."""
    import re
    
    def safe_id(node_id: str) -> str:
        return re.sub(r'[^A-Z0-9_]', '_', str(node_id).upper())
    
    def shape(node: dict) -> str:
        nid = safe_id(node.get("id", ""))
        name = node.get("name", "").replace('"', "'")
        ntype = node.get("type", "service").lower()
        if ntype == "database":
            return f'  {nid}[("{name}")]'
        elif ntype == "queue":
            return f'  {nid}[/"{name}"/]'
        elif ntype == "cache":
            return f'  {nid}{{"{name}"}}'
        elif ntype == "external" or node.get("zone") == "external":
            return f'  {nid}(["{name}"])'
        elif ntype == "client":
            return f'  {nid}>"{name}"]'
        else:
            return f'  {nid}["{name}"]'
    
    lines = ["graph TB"]
    
    # Group by zone
    zones = {}
    for node in architecture.get("nodes", []):
        zone = node.get("zone", "core")
        zones.setdefault(zone, []).append(node)
    
    # Render zones as subgraphs
    for zone_name, nodes in zones.items():
        lines.append(f'  subgraph {zone_name.upper()}["{zone_name.title()} Zone"]')
        for node in nodes:
            lines.append(f"  {shape(node).strip()}")
        lines.append("  end")
    
    # Render edges
    for edge in architecture.get("edges", []):
        src = safe_id(edge.get("source", ""))
        tgt = safe_id(edge.get("target", ""))
        proto = edge.get("protocol", "")
        if proto:
            lines.append(f'  {src} -->|{proto}| {tgt}')
        else:
            lines.append(f"  {src} --> {tgt}")
    
    # Style nodes by type
    lines.append("")
    lines.append("  classDef database fill:#fdf6e3,stroke:#b58900,stroke-width:2px")
    lines.append("  classDef external fill:#eee8d5,stroke:#586e75,stroke-width:2px")
    lines.append("  classDef service fill:#e8f4ff,stroke:#268bd2,stroke-width:2px")
    
    for node in architecture.get("nodes", []):
        nid = safe_id(node.get("id", ""))
        ntype = node.get("type", "service").lower()
        if ntype == "database":
            lines.append(f"  class {nid} database")
        elif ntype == "external" or node.get("zone") == "external":
            lines.append(f"  class {nid} external")
        else:
            lines.append(f"  class {nid} service")
    
    return "\n".join(lines)

def export_sprint_plan_markdown(sprint_plan: dict, jira_tickets: list = None) -> str:
    if not sprint_plan:
        return "# Sprint Plan\n\n_No sprint plan generated yet._"
    lines = [f"# Sprint Plan\n",
             f"**Project:** {sprint_plan.get('project','')}",
             f"**Duration:** {sprint_plan.get('sprint_duration','')}\n"]

    if jira_tickets:
        lines.append(f"## Jira Tickets Created ({len(jira_tickets)})\n")
        for t in jira_tickets:
            lines.append(f"- `{t}`")
        lines.append("")

    for epic in sprint_plan.get("epics", []):
        lines.append(f"## Epic: {epic.get('title','')}")
        lines.append(f"_{epic.get('description','')}_\n")
        for s in epic.get("stories", []):
            lines.append(f"### {s.get('story_id','')} — {s.get('title','')}")
            lines.append(f"**Story Points:** {s.get('story_points','?')}\n")
            lines.append(f"{s.get('description','')}\n")
            if s.get("acceptance_criteria"):
                lines.append("**Acceptance Criteria:**")
                for ac in s["acceptance_criteria"]:
                    lines.append(f"- {ac}")
                lines.append("")
    return "\n".join(lines)


def export_impact_markdown(impact: dict) -> str:
    if not impact:
        return "# Impact Report\n\n_No impact analysis yet._"
    lines = ["# Impact Analysis Report\n"]
    risk = impact.get("risk_assessment", {})
    lines.append(f"**Risk Level:** {risk.get('risk_level','').upper()}")
    lines.append(f"**Recommendation:** {risk.get('recommendation','')}\n")

    if risk.get("breaking_changes"):
        lines.append("## Breaking Changes\n")
        for b in risk["breaking_changes"]:
            lines.append(f"- {b}")
        lines.append("")

    if impact.get("affected_repos"):
        lines.append("## Affected Repos\n")
        for r in impact["affected_repos"]:
            lines.append(f"- `{r}`")
        lines.append("")

    if impact.get("affected_files"):
        lines.append("## Affected Files\n")
        for f in impact["affected_files"]:
            lines.append(f"- `{f.get('file_path','')}` (repo: `{f.get('repo_name','')}`, "
                        f"score: {f.get('relevance_score','')})")
        lines.append("")
    return "\n".join(lines)